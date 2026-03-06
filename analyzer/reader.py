"""Excel読取モジュール。

openpyxlを使用して.xlsx/.xlsmファイルを読み取り、
xlrdを使用して.xlsファイルを読み取る。
各セルの値と削除線情報を抽出する。
リッチテキスト（部分削除線）にも対応し、削除線のない部分のみを保持する。
"""

import logging
from dataclasses import dataclass
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

logger = logging.getLogger('analyzer')


@dataclass
class CellData:
    """単元格原始数据，包含值和删除线格式信息。"""
    value: str | None
    is_strikethrough: bool


@dataclass
class SheetData:
    """工作表原始数据，包含sheet名称和所有行的单元格数据。"""
    name: str
    rows: list[list[CellData]]


def _extract_cell(cell_data, cell_rich) -> CellData:
    """从单元格提取值，处理富文本中的部分删除线。

    对于富文本单元格，只保留没有删除线的文本片段。
    当TextBlock的strike属性为None时，继承单元格级别的字体删除线属性。
    对于普通单元格，检查整个单元格的字体是否有删除线。
    同时检测对角叉（diagonalUp + diagonalDown）作为删除标记。

    Args:
        cell_data: data_only=True 读取的单元格（含计算值）
        cell_rich: rich_text=True 读取的单元格（含富文本信息）
    """
    # 检查对角叉（X形交叉线）作为删除标记
    has_diagonal_cross = (
        cell_rich.border
        and cell_rich.border.diagonalUp
        and cell_rich.border.diagonalDown
    )
    if has_diagonal_cross:
        return CellData(value=None, is_strikethrough=True)

    rich_val = cell_rich.value

    # 富文本单元格：逐段检查删除线
    if isinstance(rich_val, CellRichText):
        # 获取单元格级别的删除线属性，作为继承的默认值
        cell_strike = bool(cell_rich.font and cell_rich.font.strike)

        kept_parts = []
        has_any_strike = False
        for part in rich_val:
            if isinstance(part, TextBlock):
                # strike=True → 明确删除线
                # strike=False → 明确非删除线
                # strike=None（有font对象）→ 有独立格式覆盖，视为非删除线
                # font=None → 无格式信息，继承单元格级别
                if part.font and part.font.strike is True:
                    is_strike = True
                elif part.font and part.font.strike is False:
                    is_strike = False
                elif part.font and part.font.strike is None:
                    # 有独立font但未设strike → 非删除线（格式覆盖）
                    is_strike = False
                else:
                    # part.font is None → 无格式信息，继承单元格级别
                    is_strike = cell_strike

                if is_strike:
                    has_any_strike = True
                else:
                    kept_parts.append(part.text)
            elif isinstance(part, str):
                # 纯文本片段（无font对象），继承单元格级别的删除线属性
                if cell_strike:
                    has_any_strike = True
                else:
                    kept_parts.append(part)

        text = "".join(kept_parts)
        # 如果全部都是删除线，标记为 strikethrough
        if not kept_parts and has_any_strike:
            return CellData(value=None, is_strikethrough=True)
        return CellData(
            value=text if text else None,
            is_strikethrough=False,
        )

    # 普通单元格：检查整个单元格级别的删除线
    is_strike = bool(cell_rich.font and cell_rich.font.strike)
    value = str(cell_data.value) if cell_data.value is not None else None
    return CellData(value=value, is_strikethrough=is_strike)


def _read_xls(file_path: Path) -> list[SheetData]:
    """xlrdを使用して.xlsファイルを読み取る。

    rich_text_runlist_mapを使用して部分削除線を検出し、
    削除線のない部分のみを保持する。

    Args:
        file_path: .xlsファイルのパス

    Returns:
        SheetDataのリスト
    """
    wb = xlrd.open_workbook(str(file_path), formatting_info=True)
    sheets = []
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        rows = []
        for ri in range(ws.nrows):
            cells = []
            for ci in range(ws.ncols):
                cell = ws.cell(ri, ci)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    cells.append(CellData(value=None, is_strikethrough=False))
                    continue

                # 値を文字列に変換
                raw = cell.value
                if cell.ctype == xlrd.XL_CELL_NUMBER:
                    value = str(int(raw)) if raw == int(raw) else str(raw)
                else:
                    value = str(raw) if raw != '' else None

                # セルレベルの削除線
                try:
                    xf_idx = ws.cell_xf_index(ri, ci)
                    xf = wb.xf_list[xf_idx]
                    cell_font = wb.font_list[xf.font_index]
                    cell_strike = bool(cell_font.struck_out)
                except Exception:
                    cell_strike = False

                # 部分削除線チェック（rich text runs）
                runs = ws.rich_text_runlist_map.get((ri, ci))
                if runs and value:
                    # runs: [(start_char, font_idx), ...]
                    # 各ランの範囲を計算して削除線のない部分のみ保持
                    kept_parts = []
                    has_any_strike = False

                    # 最初のrunより前のテキスト（セルレベルのフォントを継承）
                    first_start = runs[0][0]
                    if first_start > 0:
                        prefix = value[:first_start]
                        if cell_strike:
                            has_any_strike = True
                        else:
                            kept_parts.append(prefix)

                    for run_idx, (run_start, font_idx) in enumerate(runs):
                        run_end = (runs[run_idx + 1][0]
                                   if run_idx + 1 < len(runs)
                                   else len(value))
                        segment = value[run_start:run_end]
                        try:
                            run_font = wb.font_list[font_idx]
                            is_strike = bool(run_font.struck_out)
                        except Exception:
                            is_strike = cell_strike
                        if is_strike:
                            has_any_strike = True
                        else:
                            kept_parts.append(segment)

                    text = "".join(kept_parts)
                    if not kept_parts and has_any_strike:
                        cells.append(CellData(value=None, is_strikethrough=True))
                    else:
                        cells.append(CellData(
                            value=text if text else None,
                            is_strikethrough=False,
                        ))
                else:
                    # 通常セル：セルレベルの削除線を使用
                    if cell_strike:
                        cells.append(CellData(value=None, is_strikethrough=True))
                    else:
                        cells.append(CellData(value=value, is_strikethrough=False))
            rows.append(cells)
        sheets.append(SheetData(name=ws.name, rows=rows))
    return sheets


def read_excel(file_path: Path) -> list[SheetData]:
    """読取Excel文件，返回所有sheet的数据（含格式信息）。

    .xls形式はxlrdで、.xlsx/.xlsm形式はopenpyxlで読み取る。

    Args:
        file_path: Excel文件的路径

    Returns:
        包含所有工作表数据的SheetData列表

    Raises:
        Exception: 当文件无法打开或读取时抛出异常
    """
    if file_path.suffix.lower() == '.xls':
        return _read_xls(file_path)

    wb_data = load_workbook(file_path, data_only=True)
    wb_rich = load_workbook(file_path, rich_text=True)

    sheets = []
    for ws_data, ws_rich in zip(wb_data.worksheets, wb_rich.worksheets):
        rows = []
        for row_data, row_rich in zip(ws_data.iter_rows(), ws_rich.iter_rows()):
            cells = []
            for cd, cr in zip(row_data, row_rich):
                cells.append(_extract_cell(cd, cr))
            rows.append(cells)
        sheets.append(SheetData(name=ws_data.title, rows=rows))
    return sheets
