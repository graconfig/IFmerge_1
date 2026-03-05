"""Excel読取モジュール。

openpyxlを使用してExcelファイルを読み取り、各セルの値と削除線情報を抽出する。
リッチテキスト（部分削除線）にも対応し、削除線のない部分のみを保持する。
.xls形式のファイルは自動的に.xlsxに変換してから処理する。
"""

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

logger = logging.getLogger('analyzer')


@dataclass
class CellData:
    """単元格原始数据，包含值和删除线格式信息。"""
    value: str | None
    is_strikethrough: bool


@dataclass
class SheetData:
    """工作表原始数据，包含sheet名称和所有行的单元格数据。"""
    name: str
    rows: list[list[CellData]]


def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    """将.xls文件转换为临时的.xlsx文件（保留完整格式）。
    
    Args:
        xls_path: .xls文件路径
        
    Returns:
        临时.xlsx文件路径
    """
    try:
        import win32com.client
    except ImportError:
        raise ImportError(
            ".xls形式のファイルを処理するには pywin32 が必要です。\n"
            "インストールしてください: pip install pywin32"
        )
    
    logger.info(f".xls形式を検出: {xls_path.name} を .xlsx に変換中...")
    
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False  # 禁用警告对话框
        
        temp_dir = _get_temp_dir(xls_path)
        temp_path = temp_dir / f"{xls_path.stem}_temp.xlsx"
        
        # 打开并转换
        wb = excel.Workbooks.Open(str(xls_path.absolute()))
        wb.SaveAs(str(temp_path.absolute()), FileFormat=51)  # 51 = xlsx
        wb.Close(SaveChanges=False)
        
        logger.info(f"変換完了: {temp_path}")
        return temp_path
        
    finally:
        # 确保Excel进程正确关闭
        if excel:
            excel.Quit()

def _extract_cell(cell_data, cell_rich) -> CellData:
    """从单元格提取值，处理富文本中的部分删除线。

    对于富文本单元格，只保留没有删除线的文本片段。
    当TextBlock的strike属性为None时，继承单元格级别的字体删除线属性。
    对于普通单元格，检查整个单元格的字体是否有删除线。
    同时检测对角叉（diagonalUp + diagonalDown）作为删除标记。

    Args:
        cell_data: data_only=True 读取的单元格（含计算值）
        cell_rich: rich_text=True 读取的单元格（含富文本信息）
    """
    # 检查对角叉（X形交叉线）作为删除标记
    has_diagonal_cross = (
        cell_rich.border
        and cell_rich.border.diagonalUp
        and cell_rich.border.diagonalDown
    )
    if has_diagonal_cross:
        return CellData(value=None, is_strikethrough=True)

    rich_val = cell_rich.value

    # 富文本单元格：逐段检查删除线
    if isinstance(rich_val, CellRichText):
        # 获取单元格级别的删除线属性，作为继承的默认值
        cell_strike = bool(cell_rich.font and cell_rich.font.strike)

        kept_parts = []
        has_any_strike = False
        for part in rich_val:
            if isinstance(part, TextBlock):
                # strike=True → 明确删除线
                # strike=False → 明确非删除线
                # strike=None（有font对象）→ 有独立格式覆盖，视为非删除线
                # font=None → 无格式信息，继承单元格级别
                if part.font and part.font.strike is True:
                    is_strike = True
                elif part.font and part.font.strike is False:
                    is_strike = False
                elif part.font and part.font.strike is None:
                    # 有独立font但未设strike → 非删除线（格式覆盖）
                    is_strike = False
                else:
                    # part.font is None → 无格式信息，继承单元格级别
                    is_strike = cell_strike

                if is_strike:
                    has_any_strike = True
                else:
                    kept_parts.append(part.text)
            elif isinstance(part, str):
                # 纯文本片段（无font对象），继承单元格级别的删除线属性
                if cell_strike:
                    has_any_strike = True
                else:
                    kept_parts.append(part)

        text = "".join(kept_parts)
        # 如果全部都是删除线，标记为 strikethrough
        if not kept_parts and has_any_strike:
            return CellData(value=None, is_strikethrough=True)
        return CellData(
            value=text if text else None,
            is_strikethrough=False,
        )

    # 普通单元格：检查整个单元格级别的删除线
    is_strike = bool(cell_rich.font and cell_rich.font.strike)
    value = str(cell_data.value) if cell_data.value is not None else None
    return CellData(value=value, is_strikethrough=is_strike)


def read_excel(file_path: Path) -> list[SheetData]:
    """读取Excel文件，返回所有sheet的数据（含格式信息）。

    读取三次：
    - data_only=True: 获取公式的计算结果值
    - rich_text=True: 获取富文本信息（部分删除线）
    - data_only=False (无rich_text): 获取普通单元格的字体格式
    
    如果是.xls格式，会自动转换为临时的.xlsx文件后处理。

    Args:
        file_path: Excel文件的路径

    Returns:
        包含所有工作表数据的SheetData列表

    Raises:
        Exception: 当文件无法打开或读取时抛出异常
    """
    temp_file = None
    try:
        # 检查文件格式
        if file_path.suffix.lower() == '.xls':
            temp_file = _convert_xls_to_xlsx(file_path)
            file_to_read = temp_file
        else:
            file_to_read = file_path
        
        wb_data = load_workbook(file_to_read, data_only=True)
        wb_rich = load_workbook(file_to_read, rich_text=True)

        sheets = []
        for ws_data, ws_rich in zip(wb_data.worksheets, wb_rich.worksheets):
            rows = []
            for row_data, row_rich in zip(ws_data.iter_rows(), ws_rich.iter_rows()):
                cells = []
                for cd, cr in zip(row_data, row_rich):
                    cells.append(_extract_cell(cd, cr))
                rows.append(cells)
            sheets.append(SheetData(name=ws_data.title, rows=rows))
        
        return sheets
        
    finally:
        # 清理临时文件
        if temp_file and temp_file.exists():
            try:
                # temp_file.unlink()
                logger.debug(f"临時ファイルを削除: {temp_file}")
            except Exception as e:
                logger.warning(f"临時ファイルの削除に失敗: {e}")

def _get_temp_dir(input_file: Path) -> Path:
    """获取或创建temp目录（与input文件夹同级）。
    
    Args:
        input_file: 输入文件路径
        
    Returns:
        temp目录路径
    """
    # 获取input文件夹的父目录
    parent_dir = input_file.parent
    temp_dir = parent_dir / 'temp'
    
    # 创建temp目录（如果不存在）
    temp_dir.mkdir(exist_ok=True)
    
    return temp_dir