"""新フォーマット出力モジュール。

AI抽出結果と参考ファイル（本社EBS現行IF一覧）を組み合わせて、
IFマッピング定義書テンプレートに書き込む。

対象シート:
  - 表紙: if_name、日付を記入
  - 改訂履歴: 日付を記入
  - 対象IF: 参考ファイルから送受信システムを取得して記入
  - IFマッピング定義: AI抽出結果を記入
"""

import logging
import re
import shutil
from datetime import date
from pathlib import Path

import openpyxl

logger = logging.getLogger('analyzer')

# -----------------------------------------------------------------------
# 参考ファイル読み込み
# -----------------------------------------------------------------------

def _load_reference(reference_path: Path) -> list[dict]:
    """本社EBS現行IF一覧から K(定義書名), L(FROM), M(TO) を読み込む。"""
    wb = openpyxl.load_workbook(str(reference_path), data_only=True, read_only=True)
    ws = wb['現行IF一覧(EBS連携)']
    rows = []
    # Row 4 はヘッダー、Row 5 以降がデータ
    for row in ws.iter_rows(min_row=5, min_col=11, max_col=13):
        vals = [c.value for c in row]
        k, l, m = vals[0], vals[1], vals[2]
        if k:
            rows.append({'key': str(k).strip(), 'from': str(l).strip() if l else '', 'to': str(m).strip() if m else ''})
    wb.close()
    return rows


def _fuzzy_match(filename: str, reference_rows: list[dict]) -> list[dict]:
    """ファイル名で参考ファイルをファジーマッチする。

    戦略:
      1. 文書番号プレフィックス（BDN-XXX-XX-NNN形式）で完全一致
      2. ファイル名（拡張子なし）が参考ファイルのキーに含まれるか、またはその逆
    """
    stem = Path(filename).stem  # 拡張子なし

    # 文書番号を抽出（例: BDN-EPD-AR-008_2）
    doc_num_match = re.match(r'^(BDN-[A-Z]+-[A-Z]+-[\w]+)', stem, re.IGNORECASE)
    doc_num = doc_num_match.group(1).upper() if doc_num_match else None

    matched = []
    seen = set()
    for row in reference_rows:
        ref_key = row['key']
        ref_stem = Path(ref_key).stem.upper()
        match = False

        if doc_num and ref_stem.upper().startswith(doc_num):
            match = True
        elif stem.upper() in ref_stem or ref_stem in stem.upper():
            match = True

        if match:
            dedup_key = (row['from'], row['to'])
            if dedup_key not in seen:
                seen.add(dedup_key)
                matched.append(row)

    return matched


# -----------------------------------------------------------------------
# テンプレート書き込み
# -----------------------------------------------------------------------

def _fill_hyoshi(ws, if_name: str, today: date) -> None:
    """表紙シートを記入する。"""
    ws['C21'] = if_name
    ws['H23'] = today


def _fill_kaitei(ws, today: date) -> None:
    """改訂履歴シートの日付を記入する。"""
    ws['G2'] = today


def _replace_ebs(value: str) -> str:
    """値に EBS（大小写不敏感）が含まれる場合、値全体を SAP に置換する。"""
    if re.search(r'(?i)EBS', value):
        return 'SAP'
    return value


def _fill_taisho_if(ws, if_name: str, matched_rows: list[dict]) -> None:
    """対象IFシートを記入する。Row 6 以降に書き込む（サンプル行を上書き）。"""
    START_ROW = 6

    # サンプル行をクリア（Row 6, 7）
    for r in range(START_ROW, START_ROW + 10):
        for col in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
            ws[f'{col}{r}'] = None

    if not matched_rows:
        # マッチなしでも1行書く
        matched_rows = [{'from': '', 'to': ''}]

    for i, row in enumerate(matched_rows):
        r = START_ROW + i
        ws[f'B{r}'] = '別途採番予定'
        ws[f'C{r}'] = if_name
        ws[f'E{r}'] = '別途採番予定'
        ws[f'F{r}'] = _replace_ebs(row['from'])
        ws[f'G{r}'] = _replace_ebs(row['to'])


def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    """src_row のスタイルを dst_row にコピーする。"""
    import copy
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col_idx)
        dst = ws.cell(row=dst_row, column=col_idx)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.border = copy.copy(src.border)
            dst.fill = copy.copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy.copy(src.alignment)
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _fill_mapping(ws, records: list, sap_direction: str = "from") -> None:
    """IFマッピング定義シートを記入する。Row 5 以降に書き込む。
    テンプレートの最終行を超える場合は Row5 のスタイルをコピーする。

    sap_direction: "from" の場合は左側 From 区域（C~M列）に書き込む。
                   "to" の場合は右側 To 区域（S~AC列）に書き込む。
    """
    START_ROW = 5
    TEMPLATE_LAST_ROW = ws.max_row

    def _get(rec, field):
        if isinstance(rec, dict):
            return rec.get(field) or ''
        return getattr(rec, field, '') or ''

    use_to_side = sap_direction.lower() == "from"

    for i, rec in enumerate(records):
        r = START_ROW + i
        # 常にRow5のスタイルをコピーし、既存の値もクリアする
        _copy_row_style(ws, START_ROW, r)
        # 既存の値をクリア（テンプレートの'e'などのゴミデータを除去）
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=r, column=col_idx).value = None
        ws[f'B{r}'] = i + 1
        if use_to_side:
            ws[f'S{r}'] = _get(rec, 'item_name')         # 項目名
            ws[f'T{r}'] = _get(rec, 'dev_type')           # 標準/追加開発
            ws[f'U{r}'] = _get(rec, 'is_key')             # キー
            ws[f'V{r}'] = _get(rec, 'required')           # 必須/任意
            ws[f'W{r}'] = _get(rec, 'ebs_table_id')       # テーブルID
            ws[f'X{r}'] = _get(rec, 'item_id')            # 項目ID
            ws[f'Y{r}'] = _get(rec, 'data_type')          # データ型
            ws[f'Z{r}'] = _get(rec, 'digit_count')        # 桁数(全体)
            ws[f'AA{r}'] = _get(rec, 'digit_decimal')     # 桁数(小数点以下)
            ws[f'AB{r}'] = _get(rec, 'remarks')           # 備考
        else:
            ws[f'C{r}'] = _get(rec, 'item_name')         # 項目名
            ws[f'D{r}'] = _get(rec, 'dev_type')           # 標準/追加開発
            ws[f'E{r}'] = _get(rec, 'is_key')             # キー
            ws[f'F{r}'] = _get(rec, 'required')           # 必須/任意
            ws[f'G{r}'] = _get(rec, 'ebs_table_id')       # テーブルID
            ws[f'H{r}'] = _get(rec, 'item_id')            # 項目ID
            ws[f'I{r}'] = _get(rec, 'data_type')          # データ型
            ws[f'J{r}'] = _get(rec, 'digit_count')        # 桁数(全体)
            ws[f'K{r}'] = _get(rec, 'digit_decimal')      # 桁数(小数点以下)
            ws[f'L{r}'] = _get(rec, 'item_description')   # 項目説明
            ws[f'M{r}'] = _get(rec, 'remarks')            # 備考



# -----------------------------------------------------------------------
# メインエントリ
# -----------------------------------------------------------------------

def write_new_format(
    records: list[dict],
    if_name: str,
    input_filename: str,
    template_path: Path,
    reference_path: Path,
    output_dir: Path,
) -> Path:
    """新フォーマットExcelを生成して output_dir に保存する。

    Args:
        records: parse_response で得た dict のリスト
        if_name: AI が抽出した IF 名称
        input_filename: 入力ファイル名（参考ファイルとのマッチングに使用）
        template_path: IF抽出_新フォーマット.xlsx のパス
        reference_path: 本社EBS現行IF一覧.xlsx のパス
        output_dir: 出力先ディレクトリ

    Returns:
        生成したファイルのパス
    """
    today = date.today()
    stem = Path(input_filename).stem
    formatted_dir = output_dir / 'formatted'
    formatted_dir.mkdir(parents=True, exist_ok=True)
    out_path = formatted_dir / f"IF抽出_{stem}.xlsx"

    # テンプレートをコピー
    shutil.copy2(str(template_path), str(out_path))

    # openpyxl で開く（read_only=False で書き込み可能）
    try:
        wb = openpyxl.load_workbook(str(out_path))
    except Exception:
        # フィルター等のXML問題を回避して再試行
        wb = openpyxl.load_workbook(str(out_path), keep_vba=False)

    # 参考ファイルからマッチング
    try:
        ref_rows = _load_reference(reference_path)
        matched = _fuzzy_match(input_filename, ref_rows)
        logger.info(
            "参考ファイルマッチング: '%s' → %d件", input_filename, len(matched),
        )
    except Exception as e:
        logger.warning("参考ファイル読み込み失敗: %s", e)
        matched = []

    # 各シートに書き込み
    _fill_hyoshi(wb['表紙'], if_name, today)
    _fill_kaitei(wb['改訂履歴'], today)
    _fill_taisho_if(wb['対象IF'], if_name, matched)

    # SAP の方向を判定（置換後の from/to 値に SAP が含まれるかで判断）
    sap_direction = "from"
    for row in matched:
        if 'SAP' in _replace_ebs(row['from']).upper():
            sap_direction = "from"
            break
        if 'SAP' in _replace_ebs(row['to']).upper():
            sap_direction = "to"
            break

    _fill_mapping(wb['IFマッピング定義'], records, sap_direction)

    wb.save(str(out_path))
    wb.close()

    logger.info("新フォーマット出力: %s", out_path)
    return out_path
