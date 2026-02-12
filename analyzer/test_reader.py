from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from analyzer.reader import CellData, SheetData, read_excel


@pytest.fixture
def temp_dir(tmp_path):
    """Create a temporary directory for testing."""
    return tmp_path


def _create_simple_excel(path: Path, data: dict[str, list[list]]) -> Path:
    """Helper to create an Excel file with given sheet data.

    Args:
        path: Directory to create the file in
        data: Dict mapping sheet names to row data (list of lists)

    Returns:
        Path to the created Excel file
    """
    filepath = path / "test.xlsx"
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(filepath)
    return filepath


def _create_excel_with_strikethrough(path: Path) -> Path:
    """Helper to create an Excel file with strikethrough formatting."""
    filepath = path / "strike.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "normal text"
    ws["B1"] = "struck text"
    ws["B1"].font = Font(strike=True)
    ws["A2"] = "another normal"
    ws["B2"] = "also struck"
    ws["B2"].font = Font(strike=True)

    wb.save(filepath)
    return filepath


class TestReadExcel:
    """Tests for read_excel function."""

    def test_reads_single_sheet(self, temp_dir):
        """read_excel should read a single sheet with correct data."""
        filepath = _create_simple_excel(temp_dir, {
            "Sheet1": [["A1", "B1"], ["A2", "B2"]]
        })
        result = read_excel(filepath)
        assert len(result) == 1
        assert result[0].name == "Sheet1"
        assert len(result[0].rows) == 2

    def test_reads_multiple_sheets(self, temp_dir):
        """read_excel should read all sheets from the workbook."""
        filepath = _create_simple_excel(temp_dir, {
            "First": [["data1"]],
            "Second": [["data2"]],
            "Third": [["data3"]],
        })
        result = read_excel(filepath)
        assert len(result) == 3
        names = [s.name for s in result]
        assert names == ["First", "Second", "Third"]

    def test_extracts_cell_values_as_strings(self, temp_dir):
        """read_excel should convert cell values to strings (Req 1.5)."""
        filepath = _create_simple_excel(temp_dir, {
            "Sheet1": [["text", 123, 45.6]]
        })
        result = read_excel(filepath)
        row = result[0].rows[0]
        assert row[0].value == "text"
        assert row[1].value == "123"
        assert row[2].value == "45.6"

    def test_none_cells_remain_none(self, temp_dir):
        """read_excel should keep None for empty cells."""
        filepath = temp_dir / "empty_cells.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "value"
        # B1 is intentionally left empty
        ws["C1"] = "another"
        wb.save(filepath)

        result = read_excel(filepath)
        row = result[0].rows[0]
        assert row[0].value == "value"
        assert row[1].value is None
        assert row[2].value == "another"

    def test_detects_strikethrough_formatting(self, temp_dir):
        """read_excel should detect strikethrough formatting on cells (Req 1.5)."""
        filepath = _create_excel_with_strikethrough(temp_dir)
        result = read_excel(filepath)
        sheet = result[0]

        # Row 1: normal, struck
        assert sheet.rows[0][0].is_strikethrough is False
        assert sheet.rows[0][1].is_strikethrough is True
        # Row 2: normal, struck
        assert sheet.rows[1][0].is_strikethrough is False
        assert sheet.rows[1][1].is_strikethrough is True

    def test_non_strikethrough_cells_marked_false(self, temp_dir):
        """read_excel should mark non-strikethrough cells as False."""
        filepath = _create_simple_excel(temp_dir, {
            "Sheet1": [["hello", "world"]]
        })
        result = read_excel(filepath)
        for cell in result[0].rows[0]:
            assert cell.is_strikethrough is False

    def test_returns_sheet_data_objects(self, temp_dir):
        """read_excel should return SheetData instances."""
        filepath = _create_simple_excel(temp_dir, {
            "Sheet1": [["data"]]
        })
        result = read_excel(filepath)
        assert all(isinstance(s, SheetData) for s in result)

    def test_returns_cell_data_objects(self, temp_dir):
        """read_excel should return CellData instances for each cell."""
        filepath = _create_simple_excel(temp_dir, {
            "Sheet1": [["data"]]
        })
        result = read_excel(filepath)
        for row in result[0].rows:
            for cell in row:
                assert isinstance(cell, CellData)

    def test_empty_sheet_returns_no_rows(self, temp_dir):
        """read_excel should return empty rows list for an empty sheet."""
        filepath = temp_dir / "empty_sheet.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        # Don't write any data
        wb.save(filepath)

        result = read_excel(filepath)
        assert len(result) == 1
        assert result[0].name == "EmptySheet"
        assert result[0].rows == []

    def test_preserves_sheet_names(self, temp_dir):
        """read_excel should preserve original sheet names (Req 1.5)."""
        filepath = _create_simple_excel(temp_dir, {
            "インターフェース一覧": [["data"]],
            "詳細設計": [["data"]],
        })
        result = read_excel(filepath)
        names = [s.name for s in result]
        assert "インターフェース一覧" in names
        assert "詳細設計" in names

    def test_invalid_file_raises_exception(self, temp_dir):
        """read_excel should raise an exception for invalid files (Req 1.4)."""
        filepath = temp_dir / "invalid.xlsx"
        filepath.write_text("this is not an excel file")
        with pytest.raises(Exception):
            read_excel(filepath)

    def test_nonexistent_file_raises_exception(self, temp_dir):
        """read_excel should raise an exception for nonexistent files."""
        filepath = temp_dir / "nonexistent.xlsx"
        with pytest.raises(Exception):
            read_excel(filepath)
