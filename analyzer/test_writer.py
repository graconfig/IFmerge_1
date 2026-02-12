"""writer.pyのユニットテスト。

write_output_excel()関数の動作を検証する:
- 出力ディレクトリの自動作成 (Req 4.4)
- Output_Templateに準拠したカラム構造 (Req 4.2)
- 各レコードが1行として書き出されること (Req 4.3)
- タイムスタンプパターンのファイル名 (Req 4.5)
- 空レコードリストの処理
"""

import os
import re

from openpyxl import load_workbook

from analyzer.parser import InterfaceRecord
from analyzer.writer import write_output_excel


def test_creates_output_directory(tmp_path):
    """出力ディレクトリが存在しない場合に自動作成されること (Req 4.4)。"""
    output_dir = str(tmp_path / "new_output")
    assert not os.path.exists(output_dir)

    write_output_excel([], output_dir)

    assert os.path.isdir(output_dir)


def test_filename_matches_timestamp_pattern(tmp_path):
    """ファイル名がタイムスタンプパターンに一致すること (Req 4.5)。"""
    output_dir = str(tmp_path)
    filepath = write_output_excel([], output_dir)
    filename = os.path.basename(filepath)

    pattern = r'^EBS定義書_抽出結果_\d{8}_\d{6}\.xlsx$'
    assert re.match(pattern, filename), f"Filename '{filename}' does not match pattern"


def test_output_file_exists(tmp_path):
    """出力ファイルが実際に作成されること (Req 4.1)。"""
    output_dir = str(tmp_path)
    filepath = write_output_excel([], output_dir)

    assert os.path.isfile(filepath)


def test_header_row_matches_template(tmp_path):
    """ヘッダー行がOutput_Templateのカラム構造に一致すること (Req 4.2)。"""
    output_dir = str(tmp_path)
    filepath = write_output_excel([], output_dir)

    wb = load_workbook(filepath)
    ws = wb.active
    header_values = [cell.value for cell in ws[1]]

    expected = ['No.', '文書管理番号', 'IF名', 'EBSテーブル名',
                'EBSテーブルID', '項目ID', '項目名', '桁数']
    assert header_values == expected


def test_sheet_title(tmp_path):
    """ワークシートのタイトルが「抽出結果」であること。"""
    output_dir = str(tmp_path)
    filepath = write_output_excel([], output_dir)

    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.title == "抽出結果"


def test_records_written_as_rows(tmp_path):
    """各InterfaceRecordが1行として正しく書き出されること (Req 4.3)。"""
    records = [
        InterfaceRecord(
            document_number='DOC-001',
            if_name='IF_TEST_01',
            ebs_table_name='TABLE_A',
            ebs_table_id='T001',
            item_id='ITEM_01',
            item_name='項目A',
            digit_count='10',
        ),
        InterfaceRecord(
            document_number='DOC-002',
            if_name='IF_TEST_02',
            ebs_table_name='TABLE_B',
            ebs_table_id='T002',
            item_id='ITEM_02',
            item_name='項目B',
            digit_count='20',
        ),
    ]
    output_dir = str(tmp_path)
    filepath = write_output_excel(records, output_dir)

    wb = load_workbook(filepath)
    ws = wb.active

    # ヘッダー行 + 2データ行 = 3行
    assert ws.max_row == 3

    # 1行目のデータ検証（No.=1）
    row1 = [cell.value for cell in ws[2]]
    assert row1 == [1, 'DOC-001', 'IF_TEST_01', 'TABLE_A', 'T001',
                    'ITEM_01', '項目A', '10']

    # 2行目のデータ検証（No.=2）
    row2 = [cell.value for cell in ws[3]]
    assert row2 == [2, 'DOC-002', 'IF_TEST_02', 'TABLE_B', 'T002',
                    'ITEM_02', '項目B', '20']


def test_empty_records_produces_header_only(tmp_path):
    """空のレコードリストの場合、ヘッダー行のみが書き出されること。"""
    output_dir = str(tmp_path)
    filepath = write_output_excel([], output_dir)

    wb = load_workbook(filepath)
    ws = wb.active

    assert ws.max_row == 1  # ヘッダー行のみ


def test_no_column_auto_number(tmp_path):
    """No.列が1から始まる連番であること。"""
    records = [
        InterfaceRecord(
            document_number=f'DOC-{i:03d}',
            if_name=f'IF_{i}',
            ebs_table_name='TBL',
            ebs_table_id=f'T{i}',
            item_id=f'I{i}',
            item_name=f'名前{i}',
            digit_count=str(i * 5),
        )
        for i in range(1, 6)
    ]
    output_dir = str(tmp_path)
    filepath = write_output_excel(records, output_dir)

    wb = load_workbook(filepath)
    ws = wb.active

    for row_idx in range(2, 7):
        assert ws.cell(row=row_idx, column=1).value == row_idx - 1


def test_existing_output_directory(tmp_path):
    """既存の出力ディレクトリに対してエラーが発生しないこと。"""
    output_dir = str(tmp_path / "existing")
    os.makedirs(output_dir)

    filepath = write_output_excel([], output_dir)
    assert os.path.isfile(filepath)
